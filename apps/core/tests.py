from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from apps.billing.models import Currency, Invoice, Payment
from apps.core.management.commands.seed_initial_data import Command
from apps.core.models import AuditLog, Company, SystemSetting
from apps.core.services import build_setup_wizard, reset_company_operational_data
from apps.client_portal.models import CustomerPortalAccount
from apps.customers.models import Customer, Node
from apps.customers.services import import_customers, import_nodes
from apps.services_app.models import CustomerService, ServicePlan


class SeedInitialDataTests(TestCase):
    def test_seed_command_creates_company_settings_and_currency(self):
        Command().handle()
        self.assertTrue(Company.objects.filter(name="QvaTel").exists())
        company = Company.objects.get(name="QvaTel")
        self.assertTrue(SystemSetting.objects.filter(company=company).exists())
        self.assertTrue(Currency.objects.filter(code="USD").exists())

    def test_customer_import_accepts_alias_headers(self):
        company = Company.objects.create(name="QvaTel Import")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nombre", "telefono", "nodo", "estado", "tipo_cliente", "observaciones"])
        sheet.append(["Cliente Alias", "5551234", "Nodo Demo", "activo", "empresarial", "Importado desde alias"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "clientes.xlsx"

        result = import_customers(company, buffer)

        self.assertEqual(result["created"], 1)

    def test_customer_import_maps_payment_day_and_payment_method(self):
        company = Company.objects.create(name="QvaTel Import Payments")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nombre", "telefono", "dia_de_pago", "tipo_de_pago", "notas"])
        sheet.append(["Cliente Cobro", "5557777", "12", "paypal", "Pago portal"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "clientes_pago.xlsx"

        result = import_customers(company, buffer)

        self.assertEqual(result["created"], 1)
        customer = Customer.objects.get(company=company, full_name="Cliente Cobro")
        self.assertEqual(customer.payment_day, 12)
        self.assertEqual(customer.preferred_payment_method, Customer.PreferredPaymentMethod.PAYPAL)

    def test_customer_import_assigns_service_plan(self):
        company = Company.objects.create(name="QvaTel Import Plan")
        ServicePlan.objects.create(company=company, name="Plan 20M")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nombre", "telefono", "plan"])
        sheet.append(["Cliente Plan", "5559999", "Plan 20M"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "clientes_plan.xlsx"

        result = import_customers(company, buffer)

        self.assertEqual(result["created"], 1)
        customer = Customer.objects.get(company=company, full_name="Cliente Plan")
        self.assertTrue(customer.services.filter(plan__name="Plan 20M").exists())

    def test_node_import_accepts_alias_headers(self):
        company = Company.objects.create(name="QvaTel Nodes")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nombre", "zona", "codigo", "descripcion"])
        sheet.append(["Nodo Oeste", "Oeste", "NODO-OES-01", "Cobertura occidental"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "nodos.xlsx"

        result = import_nodes(company, buffer)

        self.assertEqual(result["created"], 1)
        node = Node.objects.get(company=company, name="Nodo Oeste")
        self.assertEqual(node.zone, "Oeste")
        self.assertEqual(node.code, "NODO-OES-01")

    def test_node_import_restores_soft_deleted_node(self):
        company = Company.objects.create(name="QvaTel Nodes Restore")
        node = Node.objects.create(company=company, name="Nodo Borrado", is_deleted=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["name", "zone", "code", "description"])
        sheet.append(["Nodo Borrado", "Centro", "NODO-001", "Nodo restaurado"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "nodos_restore.xlsx"

        result = import_nodes(company, buffer)

        self.assertEqual(result["updated"], 1)
        node.refresh_from_db()
        self.assertFalse(node.is_deleted)
        self.assertEqual(node.zone, "Centro")

    def test_setup_wizard_marks_steps_as_completed_from_existing_data(self):
        company = Company.objects.create(name="QvaTel Wizard")
        get_user_model().objects.create_superuser("wizard-admin", "wizard-admin@example.com", "secret123")
        get_user_model().objects.create_user("wizard-user", password="secret123")
        company.legal_name = "QvaTel Wizard LLC"
        company.phone = "555-0001"
        company.save(update_fields=["legal_name", "phone", "updated_at"])
        currency = Currency.objects.create(code="USDX", name="USDX", symbol="$")
        node = Node.objects.create(company=company, name="Nodo Centro")
        plan = ServicePlan.objects.create(company=company, name="Plan 30M")
        customer = Customer.objects.create(company=company, full_name="Cliente Wizard", node=node)
        service = CustomerService.objects.create(
            company=company,
            customer=customer,
            plan=plan,
            start_date=date(2026, 3, 1),
            monthly_price=Decimal("30.00"),
        )
        Invoice.objects.create(
            company=company,
            customer=customer,
            invoice_number="QVT-000001",
            due_date=date(2026, 3, 31),
            subtotal=Decimal("30.00"),
            total_amount=Decimal("30.00"),
            balance_due=Decimal("30.00"),
            currency=currency,
        )

        wizard = build_setup_wizard(company)

        self.assertEqual(wizard["completed"], wizard["total"])
        self.assertTrue(all(step["done"] for step in wizard["steps"]))
        self.assertEqual(service.customer.full_name, "Cliente Wizard")

    def test_reset_company_operational_data_removes_operational_records_and_keeps_settings(self):
        company = Company.objects.create(name="QvaTel Reset")
        setting = SystemSetting.objects.create(company=company, invoice_prefix="RST", invoice_sequence=9)
        currency = Currency.objects.create(code="RST", name="Reset Dollar", symbol="$")
        node = Node.objects.create(company=company, name="Nodo Reset")
        plan = ServicePlan.objects.create(company=company, name="Plan Reset")
        customer = Customer.objects.create(company=company, full_name="Cliente Reset", node=node)
        service = CustomerService.objects.create(
            company=company,
            customer=customer,
            plan=plan,
            start_date=date(2026, 3, 1),
            monthly_price=Decimal("25.00"),
        )
        invoice = Invoice.objects.create(
            company=company,
            customer=customer,
            invoice_number="RST-000009",
            due_date=date(2026, 3, 31),
            subtotal=Decimal("25.00"),
            total_amount=Decimal("25.00"),
            balance_due=Decimal("25.00"),
            currency=currency,
        )
        Payment.objects.create(
            company=company,
            customer=customer,
            invoice=invoice,
            amount=Decimal("10.00"),
            currency=currency,
        )
        portal = CustomerPortalAccount(customer=customer, email_login="reset@example.com")
        portal.set_password("secret123")
        portal.save()

        counts = reset_company_operational_data(company)

        self.assertEqual(counts["customers"], 1)
        self.assertEqual(counts["nodes"], 1)
        self.assertEqual(counts["plans"], 1)
        self.assertEqual(counts["customer_services"], 1)
        self.assertEqual(counts["invoices"], 1)
        self.assertEqual(counts["payments"], 1)
        self.assertFalse(Customer.objects.filter(company=company).exists())
        self.assertFalse(Node.objects.filter(company=company).exists())
        self.assertFalse(ServicePlan.objects.filter(company=company).exists())
        self.assertFalse(CustomerService.objects.filter(company=company).exists())
        self.assertFalse(Invoice.objects.filter(company=company).exists())
        self.assertFalse(Payment.objects.filter(company=company).exists())
        self.assertFalse(CustomerPortalAccount.objects.filter(email_login="reset@example.com").exists())
        setting.refresh_from_db()
        self.assertEqual(setting.invoice_sequence, 1)
        company.refresh_from_db()
        self.assertEqual(company.legal_name, "")
        self.assertEqual(company.phone, "")
        self.assertTrue(AuditLog.objects.filter(company=company, action="reset_operational_data").exists())


class WizardAndResetViewTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name="QvaTel Views")
        self.superuser = get_user_model().objects.create_superuser("admin-reset", "admin@example.com", "secret123")
        self.currency = Currency.objects.create(code="VW1", name="View Dollar", symbol="$")

    def complete_setup(self):
        self.company.legal_name = "QvaTel Views LLC"
        self.company.phone = "555-1234"
        self.company.save(update_fields=["legal_name", "phone", "updated_at"])
        get_user_model().objects.create_user("operador", password="secret123")
        node = Node.objects.create(company=self.company, name="Nodo Setup")
        plan = ServicePlan.objects.create(company=self.company, name="Plan Setup")
        customer = Customer.objects.create(company=self.company, full_name="Cliente Setup", node=node)
        CustomerService.objects.create(
            company=self.company,
            customer=customer,
            plan=plan,
            start_date=date(2026, 3, 1),
            monthly_price=Decimal("20.00"),
        )
        Invoice.objects.create(
            company=self.company,
            customer=customer,
            invoice_number="VW1-000001",
            due_date=date(2026, 3, 31),
            subtotal=Decimal("20.00"),
            total_amount=Decimal("20.00"),
            balance_due=Decimal("20.00"),
            currency=self.currency,
        )

    def test_setup_wizard_view_requires_login_and_renders(self):
        response = self.client.get(reverse("setup-wizard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Asistente inicial")

        self.client.force_login(self.superuser)
        response = self.client.get(reverse("setup-wizard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Asistente inicial")

    def test_reset_view_rejects_non_superuser(self):
        user = get_user_model().objects.create_user("operator", password="secret123")
        self.client.force_login(user)
        response = self.client.get(reverse("settings-reset"))
        self.assertEqual(response.status_code, 403)

    def test_dashboard_redirects_to_wizard_when_setup_is_pending(self):
        response = self.client.get(reverse("dashboard"))
        self.assertRedirects(response, reverse("setup-wizard"))

    def test_dashboard_renders_after_setup_completion(self):
        self.complete_setup()
        self.client.force_login(self.superuser)
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Panel operativo")

    def test_reset_total_redirects_to_wizard(self):
        self.complete_setup()
        self.client.force_login(self.superuser)
        response = self.client.post(reverse("settings-reset"), {"confirmation": "RESET TOTAL"})
        self.assertRedirects(response, reverse("setup-wizard"))

    def test_login_redirects_to_setup_while_onboarding_is_pending(self):
        response = self.client.get(reverse("login"))
        self.assertRedirects(response, reverse("setup-wizard"))

    def test_login_renders_normally_after_onboarding_completion(self):
        self.complete_setup()
        self.client.logout()
        response = self.client.get(reverse("login"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "QvaTel Cobros")

    def test_reset_removes_all_users_and_wizard_returns_to_zero(self):
        get_user_model().objects.create_user("auxiliar", password="secret123")
        counts = reset_company_operational_data(self.company, actor=self.superuser)
        wizard = build_setup_wizard(self.company)

        self.assertEqual(counts["users"], 2)
        self.assertEqual(get_user_model().objects.count(), 0)
        self.assertEqual(wizard["progress"], 0)
        self.assertFalse(wizard["steps_by_slug"]["admin"]["done"])
        self.assertFalse(wizard["steps_by_slug"]["users"]["done"])
