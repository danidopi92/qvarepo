import csv
import unicodedata
from datetime import date
from io import TextIOWrapper

from openpyxl import load_workbook

from apps.customers.models import Customer, Node
from apps.services_app.models import CustomerService, ServicePlan


HEADER_ALIASES = {
    "full_name": ["full_name", "name", "nombre", "nombre_completo", "cliente", "razon_social", "razonsocial"],
    "document_id": ["document_id", "documento", "doc", "ci", "dni", "pasaporte", "nif", "tax_id"],
    "phone": ["phone", "telefono", "telefono_principal", "mobile", "movil", "celular"],
    "whatsapp": ["whatsapp", "ws", "telefono_whatsapp"],
    "email": ["email", "correo", "correo_electronico"],
    "address": ["address", "direccion", "direccion_principal"],
    "location_reference": ["location_reference", "referencia", "referencia_ubicacion"],
    "node": ["node", "nodo", "zona"],
    "assigned_ip": ["assigned_ip", "ip", "ip_asignada"],
    "service_plan": ["service_plan", "plan", "plan_principal", "nombre_plan"],
    "payment_day": ["payment_day", "dia_de_pago", "dia_pago", "dia", "fecha_cobro"],
    "preferred_payment_method": ["preferred_payment_method", "tipo_de_pago", "tipo_pago", "metodo_pago", "forma_pago"],
    "status": ["status", "estado"],
    "customer_type": ["customer_type", "tipo_cliente", "tipo"],
    "internal_notes": ["internal_notes", "observaciones", "notas", "notas_internas"],
    "tags": ["tags", "etiquetas"],
}

NODE_HEADER_ALIASES = {
    "name": ["name", "nombre", "nodo"],
    "zone": ["zone", "zona"],
    "code": ["code", "codigo", "cod"],
    "description": ["description", "descripcion", "notas", "observaciones"],
}

STATUS_MAP = {
    "active": Customer.Status.ACTIVE,
    "activo": Customer.Status.ACTIVE,
    "activa": Customer.Status.ACTIVE,
    "suspended": Customer.Status.SUSPENDED,
    "suspendido": Customer.Status.SUSPENDED,
    "suspendida": Customer.Status.SUSPENDED,
    "inactive": Customer.Status.INACTIVE,
    "inactivo": Customer.Status.INACTIVE,
    "inactiva": Customer.Status.INACTIVE,
    "lead": Customer.Status.LEAD,
    "prospecto": Customer.Status.LEAD,
}

CUSTOMER_TYPE_MAP = {
    "residential": Customer.CustomerType.RESIDENTIAL,
    "residencial": Customer.CustomerType.RESIDENTIAL,
    "business": Customer.CustomerType.BUSINESS,
    "empresarial": Customer.CustomerType.BUSINESS,
    "empresa": Customer.CustomerType.BUSINESS,
}

PAYMENT_METHOD_MAP = {
    "cash_usd": Customer.PreferredPaymentMethod.CASH_USD,
    "efectivo_usd": Customer.PreferredPaymentMethod.CASH_USD,
    "efectivo usd": Customer.PreferredPaymentMethod.CASH_USD,
    "usd": Customer.PreferredPaymentMethod.CASH_USD,
    "cash_eur": Customer.PreferredPaymentMethod.CASH_EUR,
    "efectivo_eur": Customer.PreferredPaymentMethod.CASH_EUR,
    "efectivo eur": Customer.PreferredPaymentMethod.CASH_EUR,
    "euros": Customer.PreferredPaymentMethod.CASH_EUR,
    "eur": Customer.PreferredPaymentMethod.CASH_EUR,
    "cup": Customer.PreferredPaymentMethod.CUP,
    "paypal": Customer.PreferredPaymentMethod.PAYPAL,
    "sepa_eur": Customer.PreferredPaymentMethod.SEPA_EUR,
    "transferencia_sepa_europa": Customer.PreferredPaymentMethod.SEPA_EUR,
    "transferencia sepa europa": Customer.PreferredPaymentMethod.SEPA_EUR,
    "transfer": Customer.PreferredPaymentMethod.TRANSFER,
    "transferencia": Customer.PreferredPaymentMethod.TRANSFER,
    "transferencia_bancaria": Customer.PreferredPaymentMethod.TRANSFER,
    "crypto": Customer.PreferredPaymentMethod.CRYPTO,
    "criptomonedas": Customer.PreferredPaymentMethod.CRYPTO,
    "other": Customer.PreferredPaymentMethod.OTHER,
    "otros": Customer.PreferredPaymentMethod.OTHER,
}


def normalize_text(value):
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return normalized.strip().lower().replace("-", "_").replace(" ", "_")


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def canonicalize_row(row):
    canonical = {}
    for header, value in row.items():
        normalized_header = normalize_text(header)
        for canonical_key, aliases in HEADER_ALIASES.items():
            if normalized_header in aliases:
                canonical[canonical_key] = clean_value(value)
                break
    return canonical


def canonicalize_node_row(row):
    canonical = {}
    for header, value in row.items():
        normalized_header = normalize_text(header)
        for canonical_key, aliases in NODE_HEADER_ALIASES.items():
            if normalized_header in aliases:
                canonical[canonical_key] = clean_value(value)
                break
    return canonical


def read_rows(file_obj):
    if file_obj.name.lower().endswith(".csv"):
        reader = csv.DictReader(TextIOWrapper(file_obj.file, encoding="utf-8"))
        return [dict(row) for row in reader]

    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_value(cell) for cell in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def import_customers(company, file_obj, user=None):
    results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for index, raw_row in enumerate(read_rows(file_obj), start=2):
        row = canonicalize_row(raw_row)
        full_name = row.get("full_name", "")
        if not full_name:
            results["skipped"] += 1
            results["errors"].append(f"Fila {index}: falta full_name/nombre del cliente.")
            continue

        defaults = {
            "document_id": row.get("document_id", ""),
            "phone": row.get("phone", ""),
            "whatsapp": row.get("whatsapp", ""),
            "email": row.get("email", ""),
            "address": row.get("address", ""),
            "location_reference": row.get("location_reference", ""),
            "assigned_ip": row.get("assigned_ip") or None,
            "payment_day": int(row["payment_day"]) if row.get("payment_day", "").isdigit() else None,
            "preferred_payment_method": PAYMENT_METHOD_MAP.get(normalize_text(row.get("preferred_payment_method", "")), ""),
            "status": STATUS_MAP.get(normalize_text(row.get("status", "")), Customer.Status.ACTIVE),
            "customer_type": CUSTOMER_TYPE_MAP.get(normalize_text(row.get("customer_type", "")), Customer.CustomerType.RESIDENTIAL),
            "internal_notes": row.get("internal_notes", ""),
            "tags": row.get("tags", ""),
            "created_by": user,
            "updated_by": user,
        }

        node_name = row.get("node", "")
        if node_name:
            node, _ = Node.objects.get_or_create(company=company, name=node_name, defaults={"created_by": user, "updated_by": user})
            defaults["node"] = node

        customer, created = Customer.objects.get_or_create(
            company=company,
            full_name=full_name,
            defaults=defaults,
        )

        if created:
            assign_customer_plan(customer, row.get("service_plan", ""), user)
            results["created"] += 1
            continue

        updated = False
        for field, value in defaults.items():
            if field in {"created_by"}:
                continue
            current_value = getattr(customer, field)
            if value and current_value != value:
                setattr(customer, field, value)
                updated = True
        if updated:
            customer.updated_by = user
            customer.save()
            results["updated"] += 1
        else:
            results["skipped"] += 1

        assign_customer_plan(customer, row.get("service_plan", ""), user)

    return results


def import_nodes(company, file_obj, user=None):
    results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for index, raw_row in enumerate(read_rows(file_obj), start=2):
        row = canonicalize_node_row(raw_row)
        name = row.get("name", "")
        if not name:
            results["skipped"] += 1
            results["errors"].append(f"Fila {index}: falta nombre del nodo.")
            continue

        defaults = {
            "zone": row.get("zone", ""),
            "code": row.get("code", ""),
            "description": row.get("description", ""),
            "created_by": user,
            "updated_by": user,
        }

        node = Node.objects.filter(company=company, name=name).first()

        if not node:
            node = Node.objects.create(company=company, name=name, **defaults)
            results["created"] += 1
            continue

        updated = False
        if node.is_deleted:
            node.restore()
            updated = True
        for field, value in defaults.items():
            if field == "created_by":
                continue
            current_value = getattr(node, field)
            if value and current_value != value:
                setattr(node, field, value)
                updated = True
        if updated:
            node.updated_by = user
            node.save()
            results["updated"] += 1
        else:
            results["skipped"] += 1

    return results


def assign_customer_plan(customer, plan_name, user=None):
    plan_name = clean_value(plan_name)
    if not plan_name:
        return None

    plan = ServicePlan.objects.filter(company=customer.company, name__iexact=plan_name, is_deleted=False).first()
    if not plan:
        return None

    service = customer.services.filter(is_deleted=False).order_by("created_at").first()
    if not service:
        return CustomerService.objects.create(
            company=customer.company,
            customer=customer,
            plan=plan,
            service_type=plan.service_type,
            speed_label=plan.speed_label,
            monthly_price=plan.monthly_price,
            start_date=date.today(),
            node=customer.node,
            created_by=user,
            updated_by=user,
        )

    service.plan = plan
    service.service_type = plan.service_type
    service.speed_label = service.speed_label or plan.speed_label
    service.monthly_price = plan.monthly_price
    if not service.node and customer.node:
        service.node = customer.node
    service.updated_by = user
    service.save(update_fields=["plan", "service_type", "speed_label", "monthly_price", "node", "updated_by", "updated_at"])
    return service
